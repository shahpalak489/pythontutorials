import openpyxl
import pandas as pd


''' 
openoyxl
xlsxwriter
xlrt
xlwt
xlutils
to_excel
to_csv
'''

### to_excel https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.DataFrame.to_excel.html
li = [['Alex',10],['Bob',12],['Clarke',13]]
df2 = pd.DataFrame(li, columns=["name", "age"]) 
df2.to_excel("output1.xlsx", index=False, sheet_name='data')

### 2nd way
with pd.ExcelWriter('output.xls') as writer:  # doctest: +SKIP
	# df1.to_excel(writer, sheet_name='Sheet_name_1')
	df2.to_excel(writer, sheet_name='Sheet_name_2')

df2.to_csv("output3.csv")

###3rd way
from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")


###4th way
# from xlutils.copy import copy
# w = copy('output.xls')
# w.get_sheet(0).write(0,0,"foo")
# w.save('output5.xls')

import xlwt
from datetime import datetime

style0 = xlwt.easyxf('font: name Times New Roman, color-index red, bold on',
    num_format_str='#,##0.00')
style1 = xlwt.easyxf(num_format_str='D-MMM-YY')

wb = xlwt.Workbook()
ws = wb.add_sheet('A Test Sheet')

ws.write(0, 0, 1234.56, style0)
ws.write(1, 0, datetime.now(), style1)
ws.write(2, 0, 1)
ws.write(2, 1, 1)
ws.write(2, 2, xlwt.Formula("A3+B3"))

wb.save('example.xls')